"""
Main application controller for Battery Alert Manager
Coordinates UI, database, and Telegram integration
"""

import tkinter as tk
from tkinter import messagebox, simpledialog
import logging
import sys
from datetime import datetime
from queue import Queue, Empty
from typing import Dict, List, Optional
from pathlib import Path
import threading

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from db import DatabaseManager
from ui import AlertManagerUI
from telegram_listener import TelegramListener
from import_csv import import_csv

# Configure logging
LOG_FILE = "battery_alert_manager.log"
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


class BatteryAlertManager:
    """
    Main application controller.
    Coordinates database, UI, and Telegram integration.
    """
    
    def __init__(self):
        """Initialize the application."""
        logger.info("=" * 60)
        logger.info("Battery Alert Manager Starting")
        logger.info("=" * 60)
        
        # Initialize components
        self.db = DatabaseManager()
        self.alert_queue = Queue()
        self.telegram = TelegramListener(
            self.alert_queue,
            status_callback=self._telegram_status_callback,
            otp_callback=self._telegram_otp_callback,
            password_callback=self._telegram_password_callback
        )
        
        # Create UI
        self.root = tk.Tk()
        self.ui = AlertManagerUI(
            self.root,
            add_alert_callback=self.add_alert_manual,
            import_csv_callback=self.import_csv_file,
            update_notes_callback=self.update_alert_notes,
            update_status_callback=self.update_alert_status,
            archive_callback=self.archive_alert,
            unarchive_callback=self.unarchive_alert,
            refresh_callback=self.refresh_all_views,
            export_callback=self.export_to_excel,
            telegram_config_callback=self.configure_telegram,
            mark_critical_callback=self.mark_alert_critical,
            unmark_critical_callback=self.unmark_alert_critical
        )
        
        # Auto-refresh timer
        self.auto_refresh_enabled = True
        self.auto_refresh_interval = 5000  # 5 seconds
        
        # Queue processing
        self.queue_processing_enabled = True
        
        # Initialize UI with data
        self.refresh_all_views()
        
        # Start background tasks
        self._start_background_tasks()
        
        # Set up cleanup on close
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        logger.info("Application initialized successfully")
    
    def _start_background_tasks(self):
        """Start background tasks (queue processing, auto-refresh)."""
        # Start queue processor
        self._process_queue()
        
        # Start auto-refresh
        self._auto_refresh()
    
    def _telegram_status_callback(self, status: str):
        """Callback for Telegram status updates."""
        try:
            self.ui.update_telegram_status(status)
        except:
            pass  # UI might not be ready
    
    def _telegram_otp_callback(self) -> Optional[str]:
        """
        Callback to get OTP from user when Telegram requests it.

        This method may be invoked from the Telegram listener thread, so the
        actual dialog must be executed on the Tkinter main thread.  We use a
        temporary queue to marshal the result back to the caller.
        Returns:
            OTP code entered by user, or None if cancelled
        """
        result_queue = Queue(maxsize=1)

        def ask():
            try:
                otp = simpledialog.askstring(
                    "Telegram Verification",
                    "Enter the verification code sent to your Telegram:",
                    parent=self.root
                )
                result_queue.put(otp)
            except Exception as e:
                logger.error(f"Error showing OTP dialog: {e}")
                result_queue.put(None)

        # schedule on main thread and wait for result
        self.root.after(0, ask)
        try:
            otp = result_queue.get(timeout=300)  # give 5 minutes just in case
        except Empty:
            logger.error("OTP dialog timed out")
            return None

        if otp:
            logger.info("OTP received from user")
            return otp.strip()
        else:
            logger.warning("OTP entry cancelled by user")
            return None

    def _telegram_password_callback(self) -> Optional[str]:
        """
        Callback to get 2FA password from user when Telegram requests it.

        Runs the input dialog on the main thread and returns the password or
        None if cancelled.
        Returns:
            Password entered by user, or None if cancelled
        """
        result_queue = Queue(maxsize=1)

        def ask():
            try:
                password = simpledialog.askstring(
                    "Telegram 2FA",
                    "Enter your Telegram 2FA password:",
                    parent=self.root,
                    show='*'  # Hide password input
                )
                result_queue.put(password)
            except Exception as e:
                logger.error(f"Error showing 2FA dialog: {e}")
                result_queue.put(None)

        self.root.after(0, ask)
        try:
            password = result_queue.get(timeout=300)
        except Empty:
            logger.error("2FA password dialog timed out")
            return None

        if password:
            logger.info("2FA password received from user")
            return password
        else:
            logger.warning("2FA password entry cancelled by user")
            return None
    
    def _telegram_password_callback(self) -> Optional[str]:
        """
        Callback to get 2FA password from user when Telegram requests it.
        
        Returns:
            Password entered by user, or None if cancelled
        """
        try:
            # Use tkinter's simpledialog to get password from user
            password = simpledialog.askstring(
                "Telegram 2FA",
                "Enter your Telegram 2FA password:",
                parent=self.root,
                show='*'  # Hide password input
            )
            
            if password:
                logger.info("2FA password received from user")
                return password
            else:
                logger.warning("2FA password entry cancelled by user")
                return None
                
        except Exception as e:
            logger.error(f"Error getting 2FA password from user: {e}")
            return None
    
    def _process_queue(self):
        """Process alerts from Telegram queue."""
        if not self.queue_processing_enabled:
            return
        
        try:
            # Process all available items
            processed = 0
            while processed < 10:  # Limit per cycle
                try:
                    alert_data = self.alert_queue.get_nowait()
                    
                    # Add alert to database
                    self.db.add_or_update_alert(
                        alert_data['serial_number'],
                        alert_data['fault_type'],
                        alert_data['status'],
                        alert_data.get('notes'),
                        alert_data.get('priority','Mid')
                    )
                    
                    processed += 1
                    logger.info(f"Processed Telegram alert: {alert_data['serial_number']}")
                    
                except Empty:
                    break
            
            if processed > 0:
                # Refresh UI if we processed anything
                self.refresh_all_views()
                self.ui.update_status(f"Processed {processed} alert(s) from Telegram")
        
        except Exception as e:
            logger.error(f"Error processing queue: {e}")
        
        finally:
            # Schedule next check
            if self.queue_processing_enabled:
                self.root.after(1000, self._process_queue)
    
    def _auto_refresh(self):
        """Auto-refresh dashboards."""
        if not self.auto_refresh_enabled:
            return
        
        try:
            self.refresh_all_views()
        except Exception as e:
            logger.error(f"Error in auto-refresh: {e}")
        finally:
            if self.auto_refresh_enabled:
                self.root.after(self.auto_refresh_interval, self._auto_refresh)
    
    def add_alert_manual(
        self, 
        serial_number: str, 
        fault_type: str, 
        status: str,
        notes: str = "",
        priority: str = "Mid"
    ):
        """
        Manually add an alert.
        
        Args:
            serial_number: Device serial number
            fault_type: Fault type name
            status: Alert status
            notes: Optional notes
        """
        try:
            alert_id, was_updated = self.db.add_or_update_alert(
                serial_number,
                fault_type,
                status,
                notes if notes else None,
                priority
            )
            
            action = "updated" if was_updated else "created"
            self.ui.update_status(f"Alert {action}: ID {alert_id}")
            
            messagebox.showinfo(
                "Success",
                f"Alert {action} successfully!\nAlert ID: {alert_id}"
            )
            
            self.refresh_all_views()
            
        except Exception as e:
            logger.error(f"Error adding alert: {e}")
            messagebox.showerror("Error", f"Failed to add alert: {str(e)}")
    
    def update_alert_notes(self, alert_id: int, notes: str, priority: Optional[str] = None):
        """
        Update notes (and optionally priority) for an alert via UI callback.
        
        Args:
            alert_id: Alert ID
            notes: New notes text
            priority: Optional priority value ('High','Mid','Low')
        """
        try:
            self.db.update_alert_notes(alert_id, notes, priority)
            msg = f"Notes updated for Alert ID {alert_id}"
            if priority:
                msg += f" (priority={priority})"
            self.ui.update_status(msg)
            self.refresh_all_views()
            
        except Exception as e:
            logger.error(f"Error updating notes: {e}")
            messagebox.showerror("Error", f"Failed to update notes: {str(e)}")

    def update_alert_status(self, alert_id: int, status: str):
        """
        Change status of an alert via UI callback.
        """
        try:
            self.db.update_alert_status(alert_id, status)
            self.ui.update_status(f"Status set to {status} for Alert ID {alert_id}")
            self.refresh_all_views()
        except Exception as e:
            logger.error(f"Error updating status: {e}")
            messagebox.showerror("Error", f"Failed to update status: {str(e)}")
    
    def archive_alert(self, alert_id: int, resolution: str):
        """
        Archive an alert.
        
        Args:
            alert_id: Alert ID
            resolution: Resolution description
        """
        try:
            self.db.archive_alert(alert_id, resolution)
            self.ui.update_status(f"Alert ID {alert_id} archived")
            messagebox.showinfo("Success", "Alert archived successfully!")
            self.refresh_all_views()
            
        except Exception as e:
            logger.error(f"Error archiving alert: {e}")
            messagebox.showerror("Error", f"Failed to archive alert: {str(e)}")
    
    def unarchive_alert(self, alert_id: int):
        """
        Unarchive an alert.
        
        Args:
            alert_id: Alert ID
        """
        try:
            self.db.unarchive_alert(alert_id)
            self.ui.update_status(f"Alert ID {alert_id} unarchived")
            messagebox.showinfo("Success", "Alert unarchived successfully!")
            self.refresh_all_views()
            
        except Exception as e:
            logger.error(f"Error unarchiving alert: {e}")
            messagebox.showerror("Error", f"Failed to unarchive alert: {str(e)}")
    
    def mark_alert_critical(self, alert_id: int):
        """
        Manually mark an alert as critical.
        
        Args:
            alert_id: Alert ID
        """
        try:
            self.db.mark_as_critical(alert_id)
            self.ui.update_status(f"Alert ID {alert_id} marked as critical")
            messagebox.showinfo("Success", "Alert marked as critical!")
            self.refresh_all_views()
            
        except Exception as e:
            logger.error(f"Error marking alert as critical: {e}")
            messagebox.showerror("Error", f"Failed to mark alert as critical: {str(e)}")
    
    def unmark_alert_critical(self, alert_id: int):
        """
        Remove critical marking from an alert.
        
        Args:
            alert_id: Alert ID
        """
        try:
            self.db.unmark_as_critical(alert_id)
            self.ui.update_status(f"Alert ID {alert_id} unmarked as critical")
            messagebox.showinfo("Success", "Critical marking removed!")
            self.refresh_all_views()
            
        except Exception as e:
            logger.error(f"Error unmarking alert as critical: {e}")
            messagebox.showerror("Error", f"Failed to unmark alert as critical: {str(e)}")


    def import_csv_file(self, path: str):
        """
        Import alerts from a semicolon-delimited CSV file into the database.

        This wraps the standalone `import_csv` function and refreshes the UI.
        """
        try:
            logger.info(f"Importing CSV: {path}")
            import_csv(path, dry_run=False)
            messagebox.showinfo("Import Complete", f"Finished importing data from {path}")
            self.refresh_all_views()
        except Exception as e:
            logger.error(f"Error importing CSV file: {e}")
            messagebox.showerror("Import Error", f"Failed to import CSV: {e}")
    
    def refresh_all_views(self):
        """Refresh all dashboard views."""
        try:
            # Get data
            all_alerts = self.db.get_all_alerts(include_archived=False)
            critical_alerts = self.db.get_critical_alerts()
            archived_alerts = self.db.get_archived_alerts()
            
            # Update UI
            self.ui.populate_main_dashboard(all_alerts)
            self.ui.populate_critical_alerts(critical_alerts)
            self.ui.populate_archived_alerts(archived_alerts)
            
            # Update status
            status_msg = (
                f"Refreshed - Total: {len(all_alerts)}, "
                f"Critical: {len(critical_alerts)}, "
                f"Archived: {len(archived_alerts)}"
            )
            self.ui.update_status(status_msg)
            
        except Exception as e:
            logger.error(f"Error refreshing views: {e}")
            self.ui.update_status(f"Refresh error: {str(e)}")
    
    def export_to_excel(self, file_path: str):
        """
        Export alerts to Excel file.
        
        Args:
            file_path: Output file path
        """
        try:
            # Get all alerts
            all_alerts = self.db.get_all_alerts(include_archived=True)
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Battery Alerts"
            
            # Define styles
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            critical_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            active_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            cleared_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers
            headers = [
                "Alert ID",
                "Serial Number",
                "Fault Type",
                "Priority",
                "Status",
                "Occurred At",
                "Re-occurrence Count",
                "Notes",
                "Is Archived",
                "Resolution",
                "Archived At"
            ]
            
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Data rows
            for row, alert in enumerate(all_alerts, start=2):
                is_critical = (
                    (alert['Status'] == 'ACTIVE' and alert['ReOccurrenceCount'] >= 5)
                    or alert.get('IsCritical', 0) == 1
                ) and not alert['IsArchived']
                
                data = [
                    alert['AlertId'],
                    alert['SerialNumber'],
                    alert['FaultName'],
                    alert.get('Priority', 'Mid'),
                    alert['Status'],
                    alert['OccurredAt'],
                    alert['ReOccurrenceCount'],
                    alert.get('Notes', ''),
                    "Yes" if alert['IsArchived'] else "No",
                    alert.get('Resolution', ''),
                    alert.get('ArchivedAt', '')
                ]
                
                for col, value in enumerate(data, start=1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    
                    # Apply color coding
                    if is_critical:
                        cell.fill = critical_fill
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif alert['Status'] == 'ACTIVE' and not alert['IsArchived']:
                        cell.fill = active_fill
                    elif alert['Status'] == 'CLEARED' and not alert['IsArchived']:
                        cell.fill = cleared_fill
            
            # Adjust column widths
            column_widths = [10, 20, 25, 10, 12, 22, 18, 40, 12, 40, 22]
            for col, width in enumerate(column_widths, start=1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
            
            # Add summary sheet
            summary = wb.create_sheet("Summary")
            
            # Summary data
            total_alerts = len(all_alerts)
            active_count = sum(1 for a in all_alerts if a['Status'] == 'ACTIVE' and not a['IsArchived'])
            cleared_count = sum(1 for a in all_alerts if a['Status'] == 'CLEARED' and not a['IsArchived'])
            critical_count = sum(
                1 for a in all_alerts 
                if not a['IsArchived'] and (
                    (a['Status'] == 'ACTIVE' and a['ReOccurrenceCount'] >= 5)
                    or a.get('IsCritical', 0) == 1
                )
            )
            archived_count = sum(1 for a in all_alerts if a['IsArchived'])
            high_prio = sum(1 for a in all_alerts if a.get('Priority','Mid')=='High' and not a['IsArchived'])
            mid_prio = sum(1 for a in all_alerts if a.get('Priority','Mid')=='Mid' and not a['IsArchived'])
            low_prio = sum(1 for a in all_alerts if a.get('Priority','Mid')=='Low' and not a['IsArchived'])
            
            summary_data = [
                ["Battery Alert Manager - Summary Report"],
                ["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                [""],
                ["Metric", "Count"],
                ["Total Alerts", total_alerts],
                ["Active Alerts", active_count],
                ["Cleared Alerts", cleared_count],
                ["Critical Alerts", critical_count],
                ["Archived Alerts", archived_count],
                ["High Priority", high_prio],
                ["Mid Priority", mid_prio],
                ["Low Priority", low_prio]
            ]
            
            for row_idx, row_data in enumerate(summary_data, start=1):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = summary.cell(row=row_idx, column=col_idx, value=value)
                    
                    if row_idx == 1:
                        cell.font = Font(bold=True, size=14)
                    elif row_idx == 4:
                        cell.font = header_font
                        cell.fill = header_fill
            
            # Save workbook
            wb.save(file_path)
            
            logger.info(f"Exported {total_alerts} alerts to {file_path}")
            messagebox.showinfo(
                "Export Successful",
                f"Exported {total_alerts} alerts to:\n{file_path}"
            )
            self.ui.update_status(f"Exported {total_alerts} alerts to Excel")
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            messagebox.showerror("Export Error", f"Failed to export: {str(e)}")
    
    def configure_telegram(self, config: Dict):
        """
        Configure and start Telegram listener.
        
        Args:
            config: Dictionary with api_id, api_hash, phone, channel
        """
        try:
            self.telegram.configure(
                config['api_id'],
                config['api_hash'],
                config['phone'],
                config['channel']
            )
            
            if self.telegram.start():
                messagebox.showinfo(
                    "Telegram",
                    "Telegram listener started!\n"
                    "Monitoring for alerts..."
                )
            else:
                messagebox.showwarning(
                    "Telegram",
                    "Failed to start Telegram listener.\n"
                    "Check configuration and logs."
                )
            
        except Exception as e:
            logger.error(f"Error configuring Telegram: {e}")
            messagebox.showerror("Telegram Error", f"Configuration failed: {str(e)}")
    
    def on_closing(self):
        """Handle application closing."""
        if messagebox.askokcancel("Quit", "Do you want to quit?"):
            logger.info("Application shutting down...")
            
            # Stop background tasks
            self.auto_refresh_enabled = False
            self.queue_processing_enabled = False
            
            # Stop Telegram
            if self.telegram:
                self.telegram.stop()
            
            # Close database
            if self.db:
                self.db.close()
            
            logger.info("Shutdown complete")
            self.root.destroy()
    
    def run(self):
        """Start the application main loop."""
        logger.info("Starting UI main loop")
        self.root.mainloop()


def main():
    """Main entry point."""
    try:
        app = BatteryAlertManager()
        app.run()
    except Exception as e:
        logger.critical(f"Fatal error: {e}", exc_info=True)
        messagebox.showerror(
            "Fatal Error",
            f"Application failed to start:\n{str(e)}\n\nCheck logs for details."
        )
        sys.exit(1)


if __name__ == "__main__":
    main()